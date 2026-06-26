import asyncio
import json
import uuid
import io
from contextlib import asynccontextmanager
from datetime import datetime

from fastapi import FastAPI, Depends, HTTPException, Request, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from jinja2 import Environment, FileSystemLoader
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db, init_db
from auth import hash_password, verify_password, create_token, verify_token

import os
BASE_DIR = os.path.dirname(__file__)
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    # Auto-seed on first run (skip if admin already exists)
    pool = await get_db()
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT COUNT(*) as cnt FROM admins")
        if row["cnt"] == 0:
            await _auto_seed(conn)
    yield


app = FastAPI(lifespan=lifespan)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
jinja_env = Environment(loader=FileSystemLoader(TEMPLATES_DIR), autoescape=True)

DIMENSION_LABELS = {
    "sketch_still": "素描静物",
    "sketch_portrait": "素描头像",
    "color": "色彩",
    "quick_sketch": "速写",
    "lecture": "讲课",
    "tutoring": "辅导",
    "homework": "作业点评",
}

SEED_CLASSES = [
    "高一(1)班", "高一(2)班", "高一(3)班",
    "高二(1)班", "高二(2)班", "高二(3)班",
    "高二(4)班", "高二(5)班",
    "高三集训(1)班", "高三集训(2)班", "高三集训(3)班",
    "高三集训(4)班", "高三集训(5)班", "高三集训(6)班", "高三集训(7)班",
]

SEED_TEACHERS = [
    ("张建国", 1, "素描静物,素描头像", 1), ("李明", 1, "色彩", 0), ("王芳", 1, "速写", 0),
    ("陈志强", 2, "素描静物,素描头像", 1), ("刘洋", 2, "色彩", 0), ("赵静", 2, "速写", 0),
    ("黄伟", 3, "素描静物,素描头像", 1), ("周敏", 3, "色彩", 0), ("吴磊", 3, "速写", 0),
    ("孙立", 4, "素描头像", 1), ("郑洁", 4, "色彩", 0), ("冯涛", 4, "速写", 0),
    ("朱峰", 5, "素描静物", 1), ("韩雪", 5, "色彩", 0), ("曹刚", 5, "速写", 0),
    ("许文", 6, "素描头像", 1), ("何丽", 6, "色彩", 0), ("吕强", 6, "速写", 0),
    ("施明", 7, "素描静物", 1), ("张玲", 7, "色彩", 0), ("沈兵", 7, "速写", 0),
    ("杨帆", 8, "素描头像", 1), ("姜丽", 8, "色彩", 0), ("潘勇", 8, "速写", 0),
    ("林志远", 9, "素描头像,色彩", 1), ("谢颖", 9, "速写", 0), ("唐军", 9, "素描静物", 0),
    ("罗辉", 10, "色彩,速写", 1), ("梁静", 10, "素描头像", 0), ("宋磊", 10, "素描静物", 0),
    ("彭博", 11, "素描头像", 1), ("董洁", 11, "色彩", 0), ("袁强", 11, "速写", 0),
    ("邓超", 12, "素描静物", 1), ("叶婷", 12, "色彩", 0), ("田亮", 12, "速写", 0),
    ("程浩", 13, "色彩", 1), ("郭静", 13, "素描头像", 0), ("钟伟", 13, "速写", 0),
    ("赖涛", 14, "速写,素描头像", 1), ("肖敏", 14, "色彩", 0), ("邱磊", 14, "素描静物", 0),
    ("廖峰", 15, "素描头像", 1), ("武洁", 15, "色彩", 0), ("龙强", 15, "速写", 0),
]


async def _auto_seed(conn):
    for i, name in enumerate(SEED_CLASSES):
        await conn.execute("INSERT INTO classes (name, display_order) VALUES ($1, $2)", name, i + 1)
    for name, class_id, subjects, is_head in SEED_TEACHERS:
        await conn.execute(
            "INSERT INTO teachers (name, class_id, subjects, is_head_teacher) VALUES ($1, $2, $3, $4)",
            name, class_id, subjects, is_head,
        )
    from auth import hash_password
    await conn.execute(
        "INSERT INTO admins (username, password_hash) VALUES ($1, $2)",
        "admin", hash_password("admin123"),
    )


# ─── Middleware & helpers ────────────────────────────────────────────

def get_admin_user(request: Request):
    token = request.cookies.get("admin_token")
    if not token:
        raise HTTPException(status_code=401)
    username = verify_token(token)
    if not username:
        raise HTTPException(status_code=401)
    return username


# ─── Survey Pages ────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return RedirectResponse("/survey")


@app.get("/survey", response_class=HTMLResponse)
async def survey_page(request: Request):
    db = await get_db()
    classes = [dict(r) for r in await db.execute_fetchall("SELECT * FROM classes ORDER BY display_order")]
    await db.close()
    return HTMLResponse(jinja_env.get_template("survey.html").render(request=request, classes=classes))


@app.get("/survey/done", response_class=HTMLResponse)
async def survey_done(request: Request):
    return HTMLResponse(jinja_env.get_template("done.html").render(request=request))


# ─── Survey API ──────────────────────────────────────────────────────

@app.get("/api/classes")
async def api_classes():
    db = await get_db()
    rows = await db.execute_fetchall("SELECT id, name FROM classes ORDER BY display_order")
    await db.close()
    return [{"id": r[0], "name": r[1]} for r in rows]


@app.get("/api/teachers")
async def api_teachers(class_id: int = Query(...)):
    db = await get_db()
    rows = await db.execute_fetchall(
        "SELECT id, name, subjects, is_head_teacher FROM teachers WHERE class_id=? ORDER BY is_head_teacher DESC, id",
        (class_id,),
    )
    await db.close()
    return [{"id": r[0], "name": r[1], "subjects": r[2], "is_head_teacher": bool(r[3])} for r in rows]


@app.post("/api/submit")
async def submit_survey(request: Request):
    try:
        data = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "请求数据格式错误"}, status_code=400)

    sid = uuid.uuid4().hex[:12]
    class_id = data.get("class_id")
    db = await get_db()

    # validate class_id
    cls = await db.execute_fetchall("SELECT id FROM classes WHERE id=?", (class_id,))
    if not cls:
        await db.close()
        return JSONResponse({"ok": False, "error": "无效的班级"}, status_code=400)

    await db.execute("INSERT INTO submissions (id, class_id) VALUES (?, ?)", (sid, class_id))

    # teacher ratings
    teacher_ratings = data.get("teacher_ratings", [])
    for tr in teacher_ratings:
        teacher_id = tr.get("teacher_id")
        scores = tr.get("scores", {})
        if not teacher_id:
            continue
        await db.execute("UPDATE submissions SET teacher_id=? WHERE id=?", (teacher_id, sid))
        for dim_key, label in DIMENSION_LABELS.items():
            score = scores.get(dim_key)
            if score is not None and 1 <= score <= 10:
                await db.execute(
                    "INSERT INTO teacher_ratings (submission_id, dimension, score) VALUES (?, ?, ?)",
                    (sid, label, score),
                )

    # canteen ratings
    canteen = data.get("canteen", {})
    for dim in ["饭菜口味", "服务态度"]:
        score = canteen.get(dim)
        if score is not None and 1 <= score <= 10:
            await db.execute("INSERT INTO canteen_ratings (submission_id, dimension, score) VALUES (?, ?, ?)", (sid, dim, score))

    # dormitory ratings
    dorm = data.get("dormitory", {})
    for dim in ["宿舍管理", "宿管服务"]:
        score = dorm.get(dim)
        if score is not None and 1 <= score <= 10:
            await db.execute("INSERT INTO dormitory_ratings (submission_id, dimension, score) VALUES (?, ?, ?)", (sid, dim, score))

    # suggestion
    suggestion = data.get("suggestion", "")
    if suggestion and suggestion.strip():
        await db.execute("INSERT INTO suggestions (submission_id, content) VALUES (?, ?)", (sid, suggestion.strip()))

    await db.commit()
    await db.close()
    return {"ok": True}


# ─── Admin Pages ─────────────────────────────────────────────────────

@app.get("/admin", response_class=HTMLResponse)
async def admin_index(request: Request):
    token = request.cookies.get("admin_token")
    if token and verify_token(token):
        return RedirectResponse("/admin/dashboard")
    return RedirectResponse("/admin/login")


@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    return HTMLResponse(jinja_env.get_template("admin_login.html").render(request=request))


@app.post("/api/admin/login")
async def admin_login(request: Request):
    data = await request.json()
    username = data.get("username", "")
    password = data.get("password", "")
    db = await get_db()
    row = await db.execute_fetchall("SELECT password_hash FROM admins WHERE username=?", (username,))
    await db.close()
    if not row or not verify_password(password, row[0][0]):
        return JSONResponse({"ok": False, "error": "用户名或密码错误"}, status_code=401)
    token = create_token(username)
    resp = JSONResponse({"ok": True})
    resp.set_cookie("admin_token", token, httponly=True, max_age=86400, samesite="lax")
    return resp


@app.get("/admin/dashboard", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    token = request.cookies.get("admin_token")
    if not token or not verify_token(token):
        return RedirectResponse("/admin/login")
    return HTMLResponse(jinja_env.get_template("admin_dashboard.html").render(request=request))


@app.get("/admin/logout")
async def admin_logout():
    resp = RedirectResponse("/admin/login")
    resp.delete_cookie("admin_token")
    return resp


# ─── Admin API ───────────────────────────────────────────────────────

@app.get("/api/admin/overview")
async def api_overview(_=Depends(get_admin_user)):
    db = await get_db()
    total = (await db.execute_fetchall("SELECT COUNT(*) FROM submissions"))[0][0]
    classes_count = (await db.execute_fetchall("SELECT COUNT(*) FROM classes"))[0][0]
    teachers_count = (await db.execute_fetchall("SELECT COUNT(*) FROM teachers"))[0][0]
    suggestions_count = (await db.execute_fetchall("SELECT COUNT(*) FROM suggestions WHERE content != ''"))[0][0]

    # average by dimension
    dims = []
    for key, label in DIMENSION_LABELS.items():
        row = await db.execute_fetchall(
            "SELECT AVG(score), COUNT(*) FROM teacher_ratings WHERE dimension=?", (label,)
        )
        dims.append({"name": label, "avg": round(row[0][0] or 0, 2), "count": row[0][1]})

    await db.close()
    return {
        "total_submissions": total,
        "total_classes": classes_count,
        "total_teachers": teachers_count,
        "suggestions_count": suggestions_count,
        "dimension_averages": dims,
    }


@app.get("/api/admin/teacher-ratings")
async def api_teacher_ratings(_=Depends(get_admin_user)):
    db = await get_db()
    rows = await db.execute_fetchall("""
        SELECT t.id, t.name, c.name as class_name, t.subjects, t.is_head_teacher,
               COUNT(DISTINCT tr.submission_id) as count
        FROM teachers t
        JOIN classes c ON t.class_id = c.id
        LEFT JOIN submissions s ON s.teacher_id = t.id
        LEFT JOIN teacher_ratings tr ON tr.submission_id = s.id
        GROUP BY t.id
        ORDER BY c.display_order, t.is_head_teacher DESC, t.id
    """)

    result = []
    for r in rows:
        tid = r[0]
        dims = await db.execute_fetchall("""
            SELECT dimension, AVG(score)
            FROM teacher_ratings tr
            JOIN submissions s ON tr.submission_id = s.id
            WHERE s.teacher_id = ?
            GROUP BY dimension
        """, (tid,))
        avg_scores = {d[0]: round(d[1], 2) for d in dims}
        result.append({
            "id": tid,
            "name": r[1],
            "class_name": r[2],
            "subjects": r[3],
            "is_head_teacher": bool(r[4]),
            "count": r[5],
            "averages": avg_scores,
        })

    await db.close()
    return result


@app.get("/api/admin/canteen-ratings")
async def api_canteen_ratings(_=Depends(get_admin_user)):
    db = await get_db()
    rows = await db.execute_fetchall("""
        SELECT dimension, ROUND(AVG(score), 2), COUNT(*)
        FROM canteen_ratings
        GROUP BY dimension
    """)
    await db.close()
    return [{"dimension": r[0], "avg": r[1], "count": r[2]} for r in rows]


@app.get("/api/admin/dormitory-ratings")
async def api_dormitory_ratings(_=Depends(get_admin_user)):
    db = await get_db()
    rows = await db.execute_fetchall("""
        SELECT dimension, ROUND(AVG(score), 2), COUNT(*)
        FROM dormitory_ratings
        GROUP BY dimension
    """)
    await db.close()
    return [{"dimension": r[0], "avg": r[1], "count": r[2]} for r in rows]


@app.get("/api/admin/suggestions")
async def api_suggestions(_=Depends(get_admin_user)):
    db = await get_db()
    rows = await db.execute_fetchall("""
        SELECT s.id as sub_id, c.name as class_name, t.name as teacher_name,
               sug.content, s.created_at
        FROM suggestions sug
        JOIN submissions s ON sug.submission_id = s.id
        LEFT JOIN classes c ON s.class_id = c.id
        LEFT JOIN teachers t ON s.teacher_id = t.id
        ORDER BY s.created_at DESC
    """)
    await db.close()
    return [{"submission_id": r[0], "class": r[1] or "-", "teacher": r[2] or "未选", "content": r[3], "time": r[4]} for r in rows]


@app.get("/api/admin/export")
async def api_export(_=Depends(get_admin_user)):
    db = await get_db()
    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

    # ── Sheet 1: 教师评价汇总 ──
    ws1 = wb.active
    ws1.title = "教师评价"
    headers1 = ["班级", "教师", "科目", "班主任", "评价人数",
                "素描静物", "素描头像", "色彩", "速写", "讲课", "辅导", "作业点评", "综合均分"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers1))

    t_rows = await db.execute_fetchall("""
        SELECT t.id, t.name, c.name, t.subjects, t.is_head_teacher,
               COUNT(DISTINCT s.id) as cnt
        FROM teachers t
        JOIN classes c ON t.class_id = c.id
        LEFT JOIN submissions s ON s.teacher_id = t.id
        GROUP BY t.id ORDER BY c.display_order, t.is_head_teacher DESC
    """)

    for ri, r in enumerate(t_rows, 2):
        tid = r[0]
        dims = await db.execute_fetchall("""
            SELECT dimension, ROUND(AVG(score), 2)
            FROM teacher_ratings tr JOIN submissions s ON tr.submission_id = s.id
            WHERE s.teacher_id = ? GROUP BY dimension
        """, (tid,))
        avg_map = {d[0]: d[1] for d in dims}
        vals = [r[2], r[1], r[3], "是" if r[4] else "否", r[5] or 0]
        total = 0
        n = 0
        for dim in ["素描静物", "素描头像", "色彩", "速写", "讲课", "辅导", "作业点评"]:
            v = avg_map.get(dim)
            vals.append(v if v is not None else "-")
            if v is not None:
                total += v
                n += 1
        vals.append(round(total / n, 2) if n else "-")
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=c, value=v)
            cell.border = thin_border
            if c >= 6:
                cell.alignment = center_align

    for c in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 14 if c >= 6 else 16

    # ── Sheet 2: 饭堂评价 ──
    ws2 = wb.create_sheet("饭堂评价")
    for c, h in enumerate(["维度", "均分", "评价人次"], 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, 3)
    c_rows = await db.execute_fetchall("SELECT dimension, ROUND(AVG(score),2), COUNT(*) FROM canteen_ratings GROUP BY dimension")
    for ri, r in enumerate(c_rows, 2):
        for c, v in enumerate(r, 1):
            cell = ws2.cell(row=ri, column=c, value=v)
            cell.border = thin_border
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    # ── Sheet 3: 宿舍评价 ──
    ws3 = wb.create_sheet("宿舍评价")
    for c, h in enumerate(["维度", "均分", "评价人次"], 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, 3)
    d_rows = await db.execute_fetchall("SELECT dimension, ROUND(AVG(score),2), COUNT(*) FROM dormitory_ratings GROUP BY dimension")
    for ri, r in enumerate(d_rows, 2):
        for c, v in enumerate(r, 1):
            cell = ws3.cell(row=ri, column=c, value=v)
            cell.border = thin_border
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 12

    # ── Sheet 4: 个人建议 ──
    ws4 = wb.create_sheet("个人建议")
    for c, h in enumerate(["提交时间", "班级", "评价教师", "建议内容"], 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, 4)
    s_rows = await db.execute_fetchall("""
        SELECT s.created_at, c.name, t.name, sug.content
        FROM suggestions sug
        JOIN submissions s ON sug.submission_id = s.id
        LEFT JOIN classes c ON s.class_id = c.id
        LEFT JOIN teachers t ON s.teacher_id = t.id
        ORDER BY s.created_at DESC
    """)
    for ri, r in enumerate(s_rows, 2):
        for c, v in enumerate(r, 1):
            cell = ws4.cell(row=ri, column=c, value=v or "-")
            cell.border = thin_border
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 12
    ws4.column_dimensions["D"].width = 50

    await db.close()

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"调查问卷汇总_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
